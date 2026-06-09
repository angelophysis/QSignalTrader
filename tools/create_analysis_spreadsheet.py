"""Generates QSignalTrader manual analysis spreadsheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

OUTPUT = r'C:\Users\angel\OneDrive\Documentos\Documentos\QSignalTrader\QSignalTrader_Analises_Manuais.xlsx'

HEADER_FILL = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Calibri', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# Conditional formatting fills
FILL_GREEN = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
FILL_RED = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
FILL_GRAY = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid')
FILL_BLUE = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
FILL_CYAN = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')

FONT_GREEN = Font(color='166534')
FONT_RED = Font(color='991B1B')
FONT_YELLOW = Font(color='854D0E')
FONT_GRAY = Font(color='6B7280')
FONT_ORANGE = Font(color='9A3412')
FONT_BLUE = Font(color='1E40AF')

SHEET_ANALISES = 'Analises'
SHEET_LISTAS = 'Listas'
SHEET_RESUMO = 'Resumo'

# ── Data lists ──
TENDENCIAS = [
    '📈 Alta Leve no curto prazo',
    '📈 Tendencia de alta forte',
    '📈 Tendencia de alta no curto prazo',
    '📈 Tendencia de alta no medio/longo prazo',
    '📈 Tendencia de alta se consolidando',
    '📈 Tendencia de alta no medio prazo com possivel reversao para alta no longo prazo',
    '📉 Alta leve no curto prazo',
    '📉 Tendencia de alta no longo prazo com correcao nos demais tempos',
    '📉 Tendencia de alta no longo prazo com correcao recente',
    '📉 Tendencia de baixa forte',
    '📉 Tendencia de baixa no medio prazo com possivel continuacao para o longo prazo',
    '📉 Tendencia de baixa no curto prazo com possivel reversao para baixa no medio prazo',
    '📉 Alta leve recente',
    '🔻 Baixa leve no curto prazo',
    '🔻 Baixa leve recente',
    '🔄 Possivel reversao de tendencia no medio prazo, ainda incerta',
    '🔄 Possivel reversao de tendencia de baixa no medio prazo',
    '🔄 Possivel reversao de tendencia para baixa no medio prazo',
    '🔄 Tendencia macro preservada com correcao no curto prazo',
    '🔄 Tendencia macro de baixa preservada com correcao no curto prazo',
    '🟡 Zona de transicao - sinais mistos entre curto e medio prazo',
    '🟡 Zona de transicao - sinais mistos',
    'ℹ️ Alta apenas no 15m - nao e suficiente para indicar tendencia',
    '❌ Nenhum timeframe em tendencia de alta',
    '❌ Nenhum timeframe em tendencia clara',
    '⚠️ Perda da tendencia de alta de medio prazo',
    '⚠️ Perda da tendencia de alta de curto prazo',
    '⚠️ Perda da tendencia de alta de longo prazo',
    '⚠️ Perda da tendencia de baixa de medio prazo',
    '⚠️ Perda da tendencia de baixa de curto prazo',
    '⚠️ Perda da tendencia de baixa de longo prazo',
]

VOLATILIDADES = [
    '🧨 Volatilidade comprimida com risco de expansao',
    '🚀 Volatilidade em expansao',
    '🌋 Volatilidade alta e ainda sustentada',
    '🧊 Volatilidade elevada com probabilidade de contracao',
    '🌫 Zona de transicao da volatilidade',
]

DECISOES = [
    '🟢 Entrada comprada favorecida',
    '🟢 Entrada comprada antecipada favorecida',
    '🔴 Entrada vendida/short favorecida',
    '🔴 Entrada short antecipada favorecida',
    '🟡 Entrada comprada apenas com risco definido',
    '🟡 Entrada short apenas com risco definido',
    '🟡 Entrada somente com confirmacao',
    '🟡 Aguardar direcao ou operar apenas com confirmacao',
    '🟠 Manter, mas evitar nova entrada comprada',
    '🟠 Manter short, mas evitar nova entrada vendida',
    '🔴 Reduzir ou sair',
    '🧨 Aguardar rompimento',
    '⚪ Ficar de fora',
    '🟡 Evitar direcional',
]

LADOS = ['Long', 'Short', 'Neutro', 'Aguardar', 'Fora']
CONFIANCAS = ['Alta', 'Media', 'Baixa', 'Indefinida']
TIPOS_ATIVO = ['Cripto', 'Acao', 'ETF', 'Indice', 'Outro']
STATUS_OPS = ['Apenas observado', 'Entrada realizada', 'Aguardando confirmacao', 'Cancelado',
              'Em andamento', 'Parcial realizada', 'Encerrado com lucro', 'Encerrado com prejuizo']
RESULTADOS = ['Ainda nao avaliado', 'Acertou direcao', 'Errou direcao',
              'Volatilidade expandiu', 'Volatilidade contraiu', 'Sinal ficou neutro']

LISTAS_MAP = {
    'A': ('Tipo de Ativo', TIPOS_ATIVO),
    'B': ('Tendencia', TENDENCIAS),
    'C': ('Volatilidade', VOLATILIDADES),
    'D': ('Decisao Operacional', DECISOES),
    'E': ('Lado', LADOS),
    'F': ('Confianca', CONFIANCAS),
    'G': ('Status da Operacao', STATUS_OPS),
    'H': ('Resultado Posterior', RESULTADOS),
}

HEADERS = [
    'Ativo', 'Tipo de Ativo', 'Preco Atual', 'Preco Resistencia', 'Preco Suporte',
    'RSI', 'ATR%', 'Score Expansao', 'Score Contracao',
    'Tendencia', 'Volatilidade', 'Decisao Operacional', 'Lado', 'Confianca',
    'Data', 'Hora', 'Comentarios', 'Status da Operacao', 'Resultado Posterior', 'Data de Revisao',
]

COL_WIDTHS = [14, 14, 12, 14, 14, 8, 8, 12, 12, 42, 34, 38, 10, 12, 12, 8, 40, 20, 20, 14]

EXAMPLE_ROWS = [
    ['BTC/USDT', 'Cripto', 104000, 108000, 100000, 61.3, 2.62, 3, 5,
     '📈 Tendencia de alta se consolidando', '🚀 Volatilidade em expansao',
     '🟢 Entrada comprada favorecida', 'Long', 'Alta',
     datetime(2026, 5, 9), datetime(2026, 5, 9, 10, 30), 'Rompimento de resistencia com volatilidade acelerando',
     'Entrada realizada', 'Ainda nao avaliado', ''],
    ['ETH/USDT', 'Cripto', 3400, 3550, 3300, 45.1, 3.49, 2, 3,
     '📉 Baixa leve no curto prazo', '🧨 Volatilidade comprimida com risco de expansao',
     '🧨 Aguardar rompimento', 'Neutro', 'Media',
     datetime(2026, 5, 9), datetime(2026, 5, 9, 10, 35), 'Range definido entre 3300-3550. Aguardando.',
     'Apenas observado', 'Ainda nao avaliado', ''],
    ['SOL/USDT', 'Cripto', 195, 210, 185, 52.0, 4.10, 4, 2,
     '🟡 Zona de transicao - sinais mistos', '🌫 Zona de transicao da volatilidade',
     '⚪ Ficar de fora', 'Fora', 'Indefinida',
     datetime(2026, 5, 9), datetime(2026, 5, 9, 10, 40), 'Sinais conflitantes entre timeframes.',
     'Apenas observado', 'Ainda nao avaliado', ''],
    ['AAPL', 'Acao', 185, 190, 180, 72.2, 2.28, 6, 2,
     '📈 Tendencia de alta forte', '🌫 Zona de transicao da volatilidade',
     '🟢 Entrada comprada favorecida', 'Long', 'Alta',
     datetime(2026, 5, 9), datetime(2026, 5, 9, 10, 45), 'Tendencia forte no diario e semanal.',
     'Entrada realizada', 'Ainda nao avaliado', ''],
    ['TSLA', 'Acao', 245, 260, 235, 69.5, 5.80, 3, 4,
     '📉 Baixa leve recente', '🧊 Volatilidade elevada com probabilidade de contracao',
     '🟠 Manter, mas evitar nova entrada comprada', 'Long', 'Media',
     datetime(2026, 5, 9), datetime(2026, 5, 9, 10, 50), 'Volatilidade contraindo. Cuidado com reversao.',
     'Em andamento', 'Ainda nao avaliado', ''],
    ['QQQ', 'ETF', 420, 430, 410, 75.3, 1.90, 7, 1,
     '📈 Tendencia de alta forte', '🚀 Volatilidade em expansao',
     '🟢 Entrada comprada favorecida', 'Long', 'Alta',
     datetime(2026, 5, 9), datetime(2026, 5, 9, 10, 55), 'ETF de tecnologia mostrando forca com volatilidade subindo.',
     'Entrada realizada', 'Ainda nao avaliado', ''],
]


def _style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(ws, num_rows, num_cols):
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=(col in (10, 11, 12, 17)))


def _add_data_validation(ws, col_idx, list_range, num_rows):
    dv = DataValidation(type='list', formula1=f'=Listas!${list_range}', allow_blank=True)
    dv.error = 'Valor invalido. Selecione uma opcao da lista.'
    dv.errorTitle = 'Erro de validacao'
    col_letter = get_column_letter(col_idx)
    dv.add(f'{col_letter}2:{col_letter}{num_rows}')
    ws.add_data_validation(dv)


def _add_conditional_formatting(ws, col_idx, num_rows):
    col_letter = get_column_letter(col_idx)
    rng = f'{col_letter}2:{col_letter}{num_rows}'

    if col_idx == 10:  # Tendencia
        for pattern, fill, font in [
            ('*📈*', FILL_GREEN, FONT_GREEN),
            ('*📉*', FILL_RED, FONT_RED), ('*🔻*', FILL_RED, FONT_RED),
            ('*🟡*', FILL_YELLOW, FONT_YELLOW), ('*🔄*', FILL_YELLOW, FONT_YELLOW),
            ('*❌*', FILL_GRAY, FONT_GRAY), ('*ℹ️*', FILL_GRAY, FONT_GRAY),
            ('*⚠️*', FILL_ORANGE, FONT_ORANGE),
        ]:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator='containsText', formula=[f'"{pattern}"'], fill=fill, font=font))

    elif col_idx == 11:  # Volatilidade
        vol_rules = [
            ('*🚀*', FILL_BLUE, FONT_BLUE), ('*expansao*', FILL_BLUE, FONT_BLUE),
            ('*🧨*', FILL_YELLOW, FONT_YELLOW), ('*comprimida*', FILL_YELLOW, FONT_YELLOW),
            ('*🌋*', FILL_ORANGE, FONT_ORANGE), ('*sustentada*', FILL_ORANGE, FONT_ORANGE),
            ('*🧊*', FILL_CYAN, Font(color='0F766E')), ('*contracao*', FILL_CYAN, Font(color='0F766E')),
            ('*🌫*', FILL_GRAY, FONT_GRAY), ('*transicao*', FILL_GRAY, FONT_GRAY),
        ]
        for pattern, fill, font in vol_rules:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator='containsText', formula=[f'"{pattern}"'], fill=fill, font=font))

    elif col_idx == 12:  # Decisao
        for pattern, fill, font in [
            ('*🟢*', FILL_GREEN, FONT_GREEN),
            ('*🔴*', FILL_RED, FONT_RED),
            ('*🟡*', FILL_YELLOW, FONT_YELLOW),
            ('*⚪*', FILL_GRAY, FONT_GRAY),
            ('*🟠*', FILL_ORANGE, FONT_ORANGE),
            ('*🧨*', FILL_YELLOW, FONT_YELLOW),
        ]:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator='containsText', formula=[f'"{pattern}"'], fill=fill, font=font))


def criar_planilha():
    wb = Workbook()

    # ── Sheet: Listas ──
    ws_listas = wb.active
    ws_listas.title = SHEET_LISTAS
    for col_idx, (col_letter, (title, items)) in enumerate(LISTAS_MAP.items(), 1):
        ws_listas.cell(row=1, column=col_idx, value=title)
        for ri, val in enumerate(items, 2):
            ws_listas.cell(row=ri, column=col_idx, value=val)
    _style_header(ws_listas, len(LISTAS_MAP))
    list_max = max(len(v[1]) for v in LISTAS_MAP.values()) + 1
    _style_body(ws_listas, list_max, len(LISTAS_MAP))
    for col in range(1, len(LISTAS_MAP) + 1):
        ws_listas.column_dimensions[get_column_letter(col)].width = 42

    # ── Sheet: Analises ──
    ws = wb.create_sheet(SHEET_ANALISES)
    num_cols = len(HEADERS)
    data_rows = 500  # generous allocation

    # Headers
    for ci, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, num_cols)

    # Example rows
    for ri, row_data in enumerate(EXAMPLE_ROWS, 2):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val)

    _style_body(ws, data_rows, num_cols)

    # Column widths
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze top row + filters
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(num_cols)}1'

    # Data validation (dropdowns from Listas sheet)
    validations = {
        2: 'A2:A6',     # Tipo de Ativo
        10: 'B2:B32',   # Tendencia
        11: 'C2:C7',    # Volatilidade
        12: 'D2:D15',   # Decisao
        13: 'E2:E6',    # Lado
        14: 'F2:F5',    # Confianca
        18: 'G2:G9',    # Status
        19: 'H2:H7',    # Resultado
    }
    for col, lst_range in validations.items():
        _add_data_validation(ws, col, lst_range, data_rows)

    # Conditional formatting
    for col in [10, 11, 12]:
        _add_conditional_formatting(ws, col, data_rows)

    # Table style
    tab = Table(displayName='AnalisesTable', ref=f'A1:{get_column_letter(num_cols)}1')
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    try:
        ws.add_table(tab)
    except Exception:
        pass

    # ── Sheet: Resumo ──
    ws_res = wb.create_sheet(SHEET_RESUMO)
    col_analises = f'Analises!A2:A{data_rows}'
    col_tend = f'Analises!J2:J{data_rows}'
    col_vol = f'Analises!K2:K{data_rows}'
    col_dec = f'Analises!L2:L{data_rows}'
    col_lado = f'Analises!M2:M{data_rows}'
    col_tipo = f'Analises!B2:B{data_rows}'

    resume_items = [
        ('Resumo do Acompanhamento', '', ''),
        ('', '', ''),
        ('Total de analises registradas', f'=COUNTA({col_analises})', ''),
    ]

    # By decision
    for d in DECISOES:
        short = d[:40]
        resume_items.append((f'Decisao: {short}', f'=COUNTIF({col_dec},"{d}")', ''))

    resume_items.append(('', '', ''))
    for d in TIPOS_ATIVO:
        resume_items.append((f'Tipo: {d}', f'=COUNTIF({col_tipo},"{d}")', ''))

    resume_items.append(('', '', ''))
    resume_items.append(('Sinais Long', f'=COUNTIF({col_lado},"Long")', ''))
    resume_items.append(('Sinais Short', f'=COUNTIF({col_lado},"Short")', ''))
    resume_items.append(('Sinais Fora / Aguardar', f'=COUNTIF({col_lado},"Aguardar")+COUNTIF({col_lado},"Fora")', ''))

    for ri, (a, b, c) in enumerate(resume_items, 1):
        ws_res.cell(row=ri, column=1, value=a).font = Font(bold=(ri == 1), size=11)
        ws_res.cell(row=ri, column=2, value=b).font = BODY_FONT
        ws_res.cell(row=ri, column=3, value=c)

    ws_res.column_dimensions['A'].width = 45
    ws_res.column_dimensions['B'].width = 12

    # ── Save ──
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == '__main__':
    path = criar_planilha()
    print(f'Planilha gerada: {path}')
